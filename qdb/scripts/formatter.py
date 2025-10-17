import arrow
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from .settings import SUBCODES

COLUMNS = "ABCDEFGHIJ"
WIDTHS = [5, 10, 5] + [20] * 7

WHITE = "FFFFFFFF"
BLACK = "00000000"
RED = "00800000"

DEFAULT_FONT = Font(name="Calibri", size=12)
H1 = Font(name="Calibri", size=14, bold=True)
H1_RED = Font(name="Calibri", size=14, bold=True, color=RED)
H2 = Font(name="Calibri", size=12, bold=True)
H2_RED = Font(name="Calibri", size=12, bold=True, color=RED)
H3 = Font(name="Calibri", size=10, bold=True)
H3_RED = Font(name="Calibri", size=10, bold=True, color=RED)
NUMBER_FORMAT = "#,##0"
PERCENT_FORMAT = "0%"

BLUE_HIGHLIGHT = PatternFill(fgColor="0099CCFF", fill_type="solid")

CENTER = Alignment(horizontal="center")
RIGHT = Alignment(horizontal="right")


def adjust_column_widths(ws: Worksheet):
    """Adjust the column widths for the worksheet.

    :param ws: The worksheet to adjust
    """
    for char, width in zip(COLUMNS, WIDTHS):
        ws.column_dimensions[char].width = width
    if ws.sheet_properties.pageSetUpPr:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False


def adjust_row_heights(ws: Worksheet, lastrow: int):
    """Adjust the row heights for the worksheet.

    :param ws: The worksheet to adjust
    :param lastrow: The last row to adjust
    """
    for rownum in range(1, lastrow):
        ws.row_dimensions[rownum].height = 20


def clear_borders(ws: Worksheet, lastrow: int):
    """Clear the borders for the worksheet.

    :param ws: The worksheet to clear
    :param lastrow: The last row to clear
    """
    s = Side(color=WHITE, border_style="medium")
    for c, r in [(col, row) for row in range(1, lastrow) for col in COLUMNS]:
        ws[f"{c}{r}"].border = Border(top=s, bottom=s, left=s, right=s)
        ws[f"{c}{r}"].font = DEFAULT_FONT


def underline_row(
    ws: Worksheet, row: int, color: str, style: str = "thick", columns: str = COLUMNS
):
    """Underline the row for the worksheet.

    :param ws: The worksheet to underline
    :param row: The row to underline
    :param color: The color of the underline
    :param style: The style of the underline
    :param columns: The columns to underline
    """
    w = Side(color=WHITE, border_style="medium")
    u = Side(color=color, border_style=style)
    for c in columns:
        ws[f"{c}{row}"].border = Border(top=w, left=w, right=w, bottom=u)


def get_total_rows(account: dict, is_lib_materials: bool) -> int:
    """Get the total number of rows for the worksheet.

    :param account: The account data
    :param is_lib_materials: Whether the account is lib materials
    :return: The total number of rows
    """
    total = 7  # start with main report header
    total += (
        len(account["faus"].keys()) * 3
    )  # add lines for FAU header, spacer, and totals
    total += sum([len(fau["subs"].keys()) for fau in account["faus"].values()])
    if not is_lib_materials:
        total += 15  # add subcode legend
    return total


def calculate_fiscal_year_remainder(month: int) -> str:
    """Calculate the fiscal year remainder.

    :param month: The month to calculate the remainder for
    :return: The fiscal year remainder as a percentage
    """
    offset = 6 if month < 7 else -6
    remainder = 1 - ((month + offset) / 12)
    return f"{remainder:.0%}"


def build_report_header(
    ws: Worksheet, month: int, month_name: str, unit: str = "Unit not specified"
):
    """Build the report header for the worksheet.

    :param ws: The worksheet to build the header for
    :param month: The month to build the header for
    :param month_name: The name of the month
    :param unit: The unit to build the header for
    """
    today = arrow.now()
    # Row 1
    ws.merge_cells("A1:F1")
    ws["A1"] = today.format("MMMM DD, YYYY")
    ws.merge_cells("G1:J1")
    ws["G1"] = f"Report for: {unit}"
    ws["G1"].font = H1_RED
    ws["G1"].alignment = RIGHT
    underline_row(ws, 1, BLACK)
    # Row 2
    ws.merge_cells("A2:J2")
    ws["A2"] = "University Library and Associated Departments: General Ledger Summary"
    ws["A2"].alignment = CENTER
    ws["A2"].font = H1_RED
    # Row 3
    ws.merge_cells("A3:J3")
    ws["A3"] = f"YTD Financial Results Through the Month Ending: {month_name}"
    ws["A3"].alignment = CENTER
    ws["A3"].font = H2_RED
    # Row 4
    ws.merge_cells("A4:J4")
    amount = calculate_fiscal_year_remainder(month)
    ws["A4"] = amount + " of the fiscal year remains."
    ws["A4"].alignment = CENTER
    ws["A4"].font = H3_RED
    underline_row(ws, 4, BLACK)


def build_table_header(ws: Worksheet, row: int):
    """Build the table header for the worksheet.

    :param ws: The worksheet to build the header for
    :param row: The row to build the header for
    """
    # first row
    ws.merge_cells(f"I{row}:J{row}")
    r1_headers = [("E", "A"), ("F", "B"), ("G", "C"), ("H", "D"), ("I", "A-B-C-D")]
    for h in r1_headers:
        cel = f"{h[0]}{row}"
        ws[cel] = h[1]
        ws[cel].alignment = CENTER
        ws[cel].font = H2_RED
    row += 1
    # second row
    ws.merge_cells(f"I{row}:J{row}")
    r2_headers = [("E", "YTD"), ("F", "YTD"), ("I", "Operating Balance")]
    for h in r2_headers:
        cel = f"{h[0]}{row}"
        ws[cel] = h[1]
        ws[cel].alignment = CENTER
        ws[cel].font = H2_RED if h[0] == "I" else H2
    underline_row(ws, row, RED, style="medium", columns="IJ")
    row += 1
    # third row
    r3_headers = [
        ("E", "Appropriation"),
        ("F", "Expense"),
        ("G", "Encumbrance"),
        ("H", "Memo Lien"),
        ("I", "Amount"),
        ("J", "Percent"),
    ]
    for h in r3_headers:
        cel = f"{h[0]}{row}"
        ws[cel] = h[1]
        ws[cel].alignment = CENTER
        ws[cel].font = H2
    underline_row(ws, row, RED)
    return row + 1


def build_table(ws: Worksheet, fau: str, data: dict, row: int):
    """Build the table for the worksheet.

    :param ws: The worksheet to build the table for
    :param fau: The FAU to build the table for
    :param data: The data to build the table for
    :param row: The row to build the table for
    """
    # FAU data
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = f"{fau} {data['fund_title']}"
    ws[f"A{row}"].font = H2
    row += 1
    # iterate through subs
    for sub, subdata in data["subs"].items():
        ws[f"C{row}"] = sub
        ws[f"D{row}"] = subdata["name"]
        ws[f"E{row}"] = subdata["Appropriation"]
        ws[f"E{row}"].number_format = NUMBER_FORMAT
        ws[f"F{row}"] = subdata["Expense"]
        ws[f"F{row}"].number_format = NUMBER_FORMAT
        ws[f"G{row}"] = subdata["Encumbrance"]
        ws[f"G{row}"].number_format = NUMBER_FORMAT
        ws[f"H{row}"] = subdata["Memo Lien"]
        ws[f"H{row}"].number_format = NUMBER_FORMAT
        ws[f"I{row}"] = subdata["Amount"]
        ws[f"I{row}"].number_format = NUMBER_FORMAT
        ws[f"J{row}"] = subdata["Percent"]
        ws[f"J{row}"].number_format = PERCENT_FORMAT
        row += 1
    underline_row(ws, row - 1, RED, style="medium", columns="EFGHIJ")
    # totals
    ws[f"E{row}"] = data["totals"]["Appropriation"]
    ws[f"E{row}"].number_format = NUMBER_FORMAT
    ws[f"F{row}"] = data["totals"]["Expense"]
    ws[f"F{row}"].number_format = NUMBER_FORMAT
    ws[f"G{row}"] = data["totals"]["Encumbrance"]
    ws[f"G{row}"].number_format = NUMBER_FORMAT
    ws[f"H{row}"] = data["totals"]["Memo Lien"]
    ws[f"H{row}"].number_format = NUMBER_FORMAT
    ws[f"I{row}"] = data["totals"]["Amount"]
    ws[f"I{row}"].number_format = NUMBER_FORMAT
    return row + 2


def build_subcode_legend(ws: Worksheet, row: int):
    """Build the subcode legend for the worksheet.

    :param ws: The worksheet to build the legend for
    :param row: The row to build the legend for
    """
    row += 1
    ws[f"A{row}"] = "Sub"
    ws[f"A{row}"].font = H3
    ws[f"A{row}"].fill = BLUE_HIGHLIGHT
    ws.merge_cells(f"B{row}:D{row}")
    ws[f"B{row}"] = "Sub Title"
    ws[f"B{row}"].font = H3
    ws[f"B{row}"].fill = BLUE_HIGHLIGHT
    ws.merge_cells(f"E{row}:H{row}")
    ws[f"E{row}"] = "Notes"
    ws[f"E{row}"].font = H3
    ws[f"E{row}"].fill = BLUE_HIGHLIGHT
    row += 1
    for key in SUBCODES.keys():
        ws[f"A{row}"] = key
        ws[f"A{row}"].font = H3
        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"] = SUBCODES[key]["title"]
        ws[f"B{row}"].font = H3
        ws.merge_cells(f"E{row}:H{row}")
        ws[f"E{row}"] = SUBCODES[key]["notes"]
        ws[f"E{row}"].font = H3
        row += 1
    return row


def generate_sheet_name(account: dict, is_lib_materials: bool) -> str:
    """Generate the sheet name for the worksheet.

    :param account: The account data
    :param is_lib_materials: Whether the account is lib materials
    :return: The sheet name
    """
    name = account["account"]
    if is_lib_materials:
        name += "-LM"
    return name


def add_sub02_tab(wb: Workbook, data: dict, tab_index: int):
    """Add the sub02 tab to the workbook.

    :param wb: The workbook to add the tab to
    :param data: The data to add the tab to
    :param tab_index: The index of the tab
    """
    # create sub02 tab
    sheetname = "Sub02 Report"
    ws = wb.create_sheet(sheetname, tab_index)
    ws.title = sheetname
    # add headers
    last_row = 7 + (4 * len(data["sub02s"]))
    clear_borders(ws, last_row)
    build_report_header(ws, data["month"], data["month_name"], unit=data["unit"])
    row = build_table_header(ws, row=5)
    # add lines for each account
    for fund in data["sub02s"]:
        ws.merge_cells(f"A{row}:J{row}")
        ws[f"A{row}"] = f"{fund['fau']} {fund['fund_title']}"
        ws[f"A{row}"].font = H2
        row += 1
        # sub02 data
        ws[f"C{row}"] = "02"
        ws[f"D{row}"] = fund["row_dict"]["name"]
        ws[f"E{row}"] = fund["row_dict"]["Appropriation"]
        ws[f"E{row}"].number_format = NUMBER_FORMAT
        ws[f"F{row}"] = fund["row_dict"]["Expense"]
        ws[f"F{row}"].number_format = NUMBER_FORMAT
        ws[f"I{row}"] = fund["row_dict"]["Amount"]
        ws[f"I{row}"].number_format = NUMBER_FORMAT
        ws[f"J{row}"] = fund["row_dict"]["Percent"]
        ws[f"J{row}"].number_format = PERCENT_FORMAT
        row += 3
    # cleanup formatting
    adjust_column_widths(ws)
    adjust_row_heights(ws, row)
    Worksheet.set_printer_settings(ws, paper_size=1, orientation="landscape")
    return row


def generate_report(data: dict, filename: str):
    """Generate the report for the workbook.

    :param data: The data to generate the report for
    :param filename: The filename to generate the report for
    """
    # setup
    wb = Workbook()
    for tab_index, account in enumerate(data["accounts"]):
        is_lib_materials = "LM" in account["cc_list"]
        sheetname = generate_sheet_name(account, is_lib_materials)
        ws = wb.create_sheet(sheetname, tab_index)
        ws.title = sheetname
        last_row = get_total_rows(account, is_lib_materials)
        clear_borders(ws, last_row)
        build_report_header(ws, data["month"], data["month_name"], unit=data["unit"])
        # add tabluar data
        row = build_table_header(ws, row=5)
        for fau, fau_data in account["faus"].items():
            row = build_table(ws, fau, fau_data, row)
        if not is_lib_materials:
            row = build_subcode_legend(ws, row)
        # cleanup and save
        adjust_column_widths(ws)
        adjust_row_heights(ws, row)
        Worksheet.set_printer_settings(ws, paper_size=1, orientation="landscape")
    if len(data["sub02s"]) > 0:
        # The sub02 data needs to go in a sheet (tab) after the accounts data.
        row = add_sub02_tab(wb, data, tab_index=len(data["accounts"]))
    del wb["Sheet"]
    wb.save(filename)
    return filename
