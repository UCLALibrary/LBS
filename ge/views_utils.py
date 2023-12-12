from functools import reduce
import logging
import zipfile
from django.db.models import Model, Q
import pandas as pd
from datetime import datetime
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from os import path
from tempfile import NamedTemporaryFile
from ge.forms import ReportForm
from ge.models import BFSImport, CDWImport, LibraryData, MTFImport
from lbs.settings import BASE_DIR


logger = logging.getLogger(__name__)


def import_excel_data(excel_file: str, model: Model) -> None:
    """Import Excel data into a table.

    Parameters:
    excel_file -- The name of the Excel file.
    model -- The Django model to use.
    """
    data = get_data_from_excel(excel_file)
    # Get the Excel -> model field mapping for this model.
    # model._meta is internal, but apparently well and permanently supported.
    model_name = model._meta.object_name
    mapping = get_mapping(model_name)
    # Clear out old data before importing.
    model.objects.all().delete()
    try:
        # Iterate through rows of data, creating and saving an object for each.
        for row in data:
            # Create initial empty object.
            obj = model()
            for excel_name, field_name in mapping.items():
                # Set the value for each field.
                setattr(obj, field_name, row[excel_name])
            obj.save()
    except KeyError as ex:
        # Add useful info and re-raise up to caller
        error_message = f"Unexpected column {str(ex)} in {excel_file} for {model_name}"
        raise KeyError(error_message) from None


def get_data_from_excel(excel_file: str) -> list[dict]:
    """Read data from an Excel file (either .xls or .xlsx).
    Only reads data from the first worksheet, which is all we need.

    Returns a list of dictionaries, one for each row of data,
    keyed by the column names in the Excel file's header row.
    """
    # keep_default_na=False: Return empty strings instead of NaN or na.
    # dtype=object: Return the actual data from Excel, not an intepretation of it.
    df = pd.read_excel(excel_file, keep_default_na=False, dtype=object)
    # Uses pandas.DataFrame.to_dict with 'records' parameter:
    # 'records' : list like [{column -> value}, … , {column -> value}]
    return df.to_dict("records")


def get_mapping(model_name: str) -> dict:
    """Return a mapping of Excel column names to model field names
    for the various imports of data.

    Parameter:
    model_name -- The name of the Django model.
    """
    maps = {
        "BFSImport": {
            "Source": "source",
            "Organization": "organization",
            "OrgNumber": "fau_org",
            "Department": "department",
            "DeptNumber": "fau_dept",
            "FundType/Source": "fund_type",
            "Fund": "fau_fund",
            "FundOld": "fund_old",
            "Purpose": "purpose",
            "Description": "description",
            "BeginningBalance": "beginning_balance",
            "Contributions": "contributions",
            "InvestmentIncome": "investment_income",
            "GiftFeePayments": "gift_fee_payments",
            "RealGl": "real_gl",
            "UnrealGL": "unreal_gl",
            "FoundationTransfers": "foundation_transfers",
            "TransfersToUniv": "transfers_to_univ",
            "Expenditures": "expenditures",
            "Transfers/Adj": "transfers_adj",
            "EndingBalance": "ending_balance",
            "Available": "available",
            "Unavailable": "unavailable",
            "Principal": "principal",
            "MarketValue": "market_value",
            "ProjectedIncome": "projected_income",
            "SortAccount": "sort_account",
            "FundSummary": "fund_summary",
            "FundMemo": "fund_memo",
            "PeriodStart": "period_start",
            "PeriodEnd": "period_end",
        },
        "CDWImport": {
            "Ledger Year Month": "ledger_year_month",
            "FYE Proc Indicator": "fye_proc_indicator",
            "Account Org Code": "fau_org",
            "Fund Group": "fund_group",
            "Account Department Code": "fau_dept",
            "Location Code": "fau_location",
            "Account Number": "fau_account",
            "Cost Center Code": "fau_cost_center",
            "Fund Number": "fau_fund",
            "Fund Title": "fau_fund_title",
            "Inception-to-Date Appropriation": "inception_to_date_appropriation",
            "Inception-to-Date Financial": "inception_to_date_financial",
            "Encum&ML": "encum_ml",
            "Operating Balance": "operating_balance",
        },
        "MTFImport": {
            "org_code": "fau_org",
            "org_title": "organization",
            "division_code": "division_code",
            "division_title": "division_title",
            "sub_division_code": "sub_division_code",
            "sub_division_title": "sub_division_title",
            "dept_code": "fau_dept",
            "dept_title": "department",
            "fund_id": "fund_id",
            "fund_nbr": "fund_nbr",
            "fund_desc": "fund_description",
            "fiscal_yr": "fiscal_year",
            "last_fiscal_month_closed": "last_fiscal_month_closed",
            "fdn_dept": "fdn_dept",
            "fdn_dept_desc": "fdn_dept_description",
            "univ_dept": "univ_dept",
            "univ_dept_desc": "univ_dept_description",
            "univ_fund_nbr": "fau_fund",
            "univ_fund_desc": "fau_fund_title",
            "use_code": "use_code",
            "use_desc": "use_description",
            "available_bal": "available_balance",
            "unavailable_bal": "unavailable_balance",
            "pending_transfer_bal": "pending_transfer_balance",
            "max_transfer_bal": "max_transfer_balance",
            "fund_purpose": "fund_purpose",
            "fund_restriction": "fund_restriction",
            "signature_ind": "signature_ind",
            "preparer_phone_num": "preparer_phone_number",
        },
        "LibraryData": {
            "ID": "original_id",
            "UnitGrande": "unit_grande",
            "Unit": "unit",
            "Home Unit/Dept": "home_unit_dept",
            "Fund Title": "fund_title",
            "Fund Type": "fund_type",
            "Reg/Fdn": "reg_fdn",
            "Fund Manager": "fund_manager",
            "UCOP/ FDN No": "ucop_fdn_no",
            "Fund No": "fau_fund_no",
            "Account": "fau_account",
            "CC": "fau_cost_center",
            "Fund": "fau_fund",
            "YTD Appropriation": "ytd_appropriation",
            "YTD Expenditure": "ytd_expenditure",
            "Commitments": "commitments",
            "Operating Balance": "operating_balance",
            "Max MTF Trf Amt": "max_mtf_trf_amt",
            "Total Balance": "total_balance",
            "MTF_Authority": "mtf_authority",
            "Total Fund Value": "total_fund_value",
            "Projected Annual Income": "projected_annual_income",
            "Fund Summary": "fund_summary",
            "Fund Purpose": "fund_purpose",
            "Notes": "notes",
            "Home Dept": "home_dept",
            "FundRestriction": "fund_restriction",
            "NewFund": "new_fund",
            "LBS Notes": "lbs_notes",
        },
    }
    return maps[model_name]


def add_funds() -> None:
    """Add funds from campus data not already in LibraryData
    for use in reports.
    """

    # Original Access query qryAddNew_2:
    # Add rows to LibraryData from CDWImport where
    # (fau_account, fau_cost_center, fau_fund) doesn't already exist.
    incoming_funds = CDWImport.objects.all().values(
        "fau_account", "fau_cost_center", "fau_fund"
    )
    current_funds = LibraryData.objects.all().values(
        "fau_account", "fau_cost_center", "fau_fund"
    )
    for f in incoming_funds:
        if f not in current_funds:
            new_fund = LibraryData(
                fau_account=f["fau_account"],
                fau_cost_center=f["fau_cost_center"],
                fau_fund=f["fau_fund"],
                new_fund="Y",
            )

            # Original Access query qryAddNew_3:
            # Match on BFSImport data to set other values in the new fund.
            # Match is only on fau_fund, which often can find multiple rows,
            # but the relevant fields appear to be the same for any given fund.
            # Take the "first" match - if any, since match is not guaranteed.
            bfs_funds = BFSImport.objects.filter(fau_fund=new_fund.fau_fund)
            if bfs_funds.exists():
                bfs_fund = bfs_funds[0]
                new_fund.fund_title = bfs_fund.description
                new_fund.fund_type = bfs_fund.fund_type
                new_fund.fund_summary = bfs_fund.fund_summary
                new_fund.fund_purpose = bfs_fund.purpose

                # Original Access query qryAddNew_4 - queryAddNew_7:
                # Normalize reg_fdn and fund_type values.
                # These are converted from Access and assume all data... matches assumptions.
                # These are dependent on a matching BFSImport row being found above.
                if "FOUNDATION" in new_fund.fund_type.upper():
                    new_fund.reg_fdn = "F"
                if "REGENTAL" in new_fund.fund_type.upper():
                    new_fund.reg_fdn = "R"
                if "ENDOWMENT" in new_fund.fund_type.upper():
                    new_fund.fund_type = "Endowment"
                if "EXPENDITURE" in new_fund.fund_type.upper():
                    new_fund.fund_type = "Current Expenditure"

            else:
                # No matching BFSImport row found, so log a message.
                logger.warning(
                    f"No matching BFS (consolidated) data found for {new_fund.fau_fund=}"
                )

            # Original Access query qryAddNew_8:
            # Clear the "new_fund" flag.
            new_fund.new_fund = "N"

            # Finally, save the new LibraryData record.
            logger.info(f"Added fund: {new_fund}")
            new_fund.save()


def update_data() -> None:
    """Update LibraryData rows to final state before report generation."""

    # Original Access query qryAAA_0Clear:
    # Set several financial values to 0 for all rows.
    cnt = LibraryData.objects.all().update(
        ytd_appropriation=0,
        ytd_expenditure=0,
        commitments=0,
        operating_balance=0,
        max_mtf_trf_amt=0,
        projected_annual_income=0,
        total_fund_value=0,
    )

    logger.info(f"qryAAA_0Clear: {cnt} updated")

    # Original Access query qryAAA_1UpdateMTF (and duplicate qryAAA_1UpdateMTF1):
    # Update relevant LibraryData rows from MTF data (first matching row only, if any).
    cnt = 0
    for ld in LibraryData.objects.all():
        if ld.ucop_fdn_no:
            mtf_rows = MTFImport.objects.filter(fund_nbr=ld.ucop_fdn_no)
            if mtf_rows.exists():
                mtf = mtf_rows[0]
                ld.fund_restriction = mtf.fund_restriction
                ld.max_mtf_trf_amt = mtf.max_transfer_balance
                ld.save()
                cnt += 1
    logger.info(f"qryAAA_1UpdateMTF: {cnt} updated")

    # Original Access query qryAAA_2ProjIncomFound:
    # Update projected annual income from BFS data (Foundation funds).
    cnt = 0
    for ld in LibraryData.objects.all():
        if ld.ucop_fdn_no:
            bfs_rows = BFSImport.objects.filter(fau_fund=ld.ucop_fdn_no, source="F")
            if bfs_rows.exists():
                bfs = bfs_rows[0]
                ld.projected_annual_income = bfs.projected_income
                ld.save()
                cnt += 1
    logger.info(f"qryAAA_2ProjIncomFound: {cnt} updated")

    # Original Access query qryAAA_2ProjIncomReg:
    # Update projected annual income from BFS data (Regental funds).
    cnt = 0
    for ld in LibraryData.objects.all():
        if ld.fau_fund:
            bfs_rows = BFSImport.objects.filter(fau_fund=ld.fau_fund, source="R")
            if bfs_rows.exists():
                bfs = bfs_rows[0]
                ld.projected_annual_income = bfs.projected_income
                ld.save()
                cnt += 1
    logger.info(f"qryAAA_2ProjIncomReg: {cnt} updated")

    # Original Access query qryAAA_3FoundTotVal:
    cnt = 0
    for ld in LibraryData.objects.filter(reg_fdn="F"):
        if ld.ucop_fdn_no:
            bfs_rows = BFSImport.objects.filter(
                fau_fund=ld.ucop_fdn_no,
                fund_type__in=("ENDOWMENT REGENTAL INCOME", "ENDOWMENT FOUNDATION"),
            )
            if bfs_rows.exists():
                bfs = bfs_rows[0]
                ld.total_fund_value = bfs.market_value
                ld.save()
                cnt += 1
    logger.info(f"qryAAA_3FoundTotVal: {cnt} updated")

    # Original Access query qryAAA_3FoundTotVal_2:
    cnt = 0
    for ld in LibraryData.objects.filter(reg_fdn="F"):
        if ld.fau_fund:
            bfs_rows = BFSImport.objects.filter(
                fau_fund=ld.fau_fund,
                fund_type__in=("ENDOWMENT REGENTAL INCOME", "ENDOWMENT FOUNDATION"),
            )
            if bfs_rows.exists():
                bfs = bfs_rows[0]
                # Seems like this will always be just bfs.available,
                # since ld.total_fund_value is set to 0 by qryAAA_0Clear,
                # and this set doesn't intersect qryAAA_3FoundTotVal...
                # so ld.total_fund_value was not updated by bfs.market_value.
                # Per LBS, total_fund_value is not really used...
                # but we'll replicate legacy logic for consistency.
                ld.total_fund_value = ld.total_fund_value + bfs.available
                ld.save()
                cnt += 1
    logger.info(f"qryAAA_3FoundTotVal_2: {cnt} updated")

    # Original Access query qryAAA_3FoundTotVal_3:
    # This currently matches no rows.
    # Per LBS, total_fund_value is not really used...
    # but we'll replicate legacy logic for consistency.
    cnt = 0
    for ld in LibraryData.objects.filter(reg_fdn="F"):
        if ld.fau_fund:
            bfs_rows = BFSImport.objects.filter(
                fau_fund=ld.fau_fund,
                source="R",
                fund_type="ENDOWMENT FOUNDATION",
            )
            if bfs_rows.exists():
                bfs = bfs_rows[0]
                ld.total_fund_value = ld.total_fund_value - bfs.unavailable
                ld.save()
                cnt += 1
    logger.info(f"qryAAA_3FoundTotVal_3: {cnt} updated")

    # Original Access query qryAAA_3RegTotVal:
    # Regental fund math is different from Foundation math above...
    cnt = 0
    for ld in LibraryData.objects.filter(reg_fdn="R"):
        if ld.fau_fund:
            bfs_rows = BFSImport.objects.filter(
                fau_fund=ld.fau_fund,
                source="U",
            )
            if bfs_rows.exists():
                bfs = bfs_rows[0]
                ld.total_fund_value = bfs.available
                ld.save()
                cnt += 1
    logger.info(f"qryAAA_3RegTotVal: {cnt} updated")

    # Original Access query qryAAA_3RegTotVal_2:
    # Regental fund math is different from Foundation math above...
    cnt = 0
    for ld in LibraryData.objects.filter(reg_fdn="R"):
        if ld.fau_fund:
            bfs_rows = BFSImport.objects.filter(
                fau_fund=ld.fau_fund,
                source="R",
            )
            if bfs_rows.exists():
                bfs = bfs_rows[0]
                ld.total_fund_value = ld.total_fund_value - bfs.unavailable
                ld.save()
                cnt += 1
    logger.info(f"qryAAA_3RegTotVal_2: {cnt} updated")

    # Original Access query qryAAA_3RegTotVal_3:
    # Regental fund math is different from Foundation math above...
    cnt = 0
    for ld in LibraryData.objects.filter(reg_fdn="R"):
        if ld.ucop_fdn_no:
            bfs_rows = BFSImport.objects.filter(
                fau_fund=ld.ucop_fdn_no,
                source="R",
            )
            if bfs_rows.exists():
                bfs = bfs_rows[0]
                ld.total_fund_value = ld.total_fund_value + bfs.market_value
                ld.save()
                cnt += 1
    logger.info(f"qryAAA_3RegTotVal_3: {cnt} updated")

    # Original Access query qryAAA_3RegTotVal_4:
    # Regental fund math is different from Foundation math above...
    cnt = 0
    for ld in LibraryData.objects.filter(reg_fdn="R"):
        if ld.fau_fund_no:
            bfs_rows = BFSImport.objects.filter(
                fau_fund=ld.fau_fund_no,
                fau_fund__gt="40000",
                source="U",
            )
            if bfs_rows.exists():
                bfs = bfs_rows[0]
                ld.total_fund_value = bfs.available
                ld.save()
                cnt += 1
    logger.info(f"qryAAA_3RegTotVal_4: {cnt} updated")

    # Original Access query qryAAA_5_5400:
    # Update LibraryData amounts from CDW data.
    cnt = 0
    for ld in LibraryData.objects.all():
        if ld.fau_fund and ld.fau_cost_center and ld.fau_account:
            cdw_rows = CDWImport.objects.filter(
                fau_fund=ld.fau_fund,
                fau_cost_center=ld.fau_cost_center,
                fau_account=ld.fau_account,
            )
            if cdw_rows.exists():
                cdw = cdw_rows[0]
                ld.ytd_appropriation = cdw.inception_to_date_appropriation
                ld.ytd_expenditure = cdw.inception_to_date_financial
                ld.commitments = cdw.encum_ml
                ld.operating_balance = cdw.operating_balance
                ld.save()
                cnt += 1
    logger.info(f"qryAAA_5_5400: {cnt} updated")

    # Original Access query qryAAA_6TotalBalance:
    # Finally, update LibraryData total balance for all rows.
    cnt = 0
    for ld in LibraryData.objects.all():
        ld.total_balance = ld.operating_balance + ld.max_mtf_trf_amt
        ld.save()
        cnt += 1
    logger.info(f"qryAAA_6TotalBalance: {cnt} updated")


def get_librarydata_results(search_type: str, search_term: str) -> list[LibraryData]:
    """Search LibraryData fields for a search_term, based on search_type.

    Parameters:
    search_type -- The type of search (fund or keyword)
    search_term -- The term to search for

    Returns a list of LibraryData objects matching the search.
    """
    if search_type == "fund":
        fields_to_search = ["fau_fund", "fau_fund_no", "ucop_fdn_no"]
    elif search_type == "keyword":
        fields_to_search = [
            "fund_manager",
            "fund_purpose",
            "fund_restriction",
            "fund_summary",
            "fund_title",
            "lbs_notes",
            "notes",
        ]
    elif search_type == "unit":
        fields_to_search = ["unit"]
    elif search_type == "new_funds":
        fields_to_search = ["fund_manager", "unit"]
    else:
        raise ValueError(f"Unsupported search type: {search_type}")

    # Search logic is different for new funds
    if search_type == "new_funds":
        # Look for empty fields, overriding search_term if supplied.
        search_term = ""
        q_list = [Q(**{field + "__exact": search_term}) for field in fields_to_search]
        # AND them all together.
        q_filter = reduce(lambda a, b: a & b, q_list)
    else:
        # Get a list of individual Q() statements looking for search_term in each field.
        # Example result: [<Q: (AND: ('field_a', 'term'))>, <Q: (AND: ('field_b', 'term'))>]
        q_list = [
            Q(**{field + "__icontains": search_term}) for field in fields_to_search
        ]
        # OR them all together.
        q_filter = reduce(lambda a, b: a | b, q_list)

    # Apply the filter to find results.
    results = LibraryData.objects.filter(q_filter).order_by("id")

    # Return results as a list of objects, rather than a queryset.
    return [item for item in results]


def sum_col(ws: Worksheet, col: str, col_top: int = 5) -> None:
    """Add a total row to an Excel spreadsheet column."""
    last_row = get_last_row(ws, col)
    ws[f"{col}{last_row + 1}"] = f"=SUM({col}{col_top}:{col}{last_row})"


def get_last_row(ws: Worksheet, col: str) -> int:
    """Get the index of the last non-empty row in an Excel spreadsheet column."""
    last_row = len(ws[col])
    # make sure we get the first non-empty row from the bottom
    last_row -= next(i for i, x in enumerate(reversed(ws[col])) if x.value is not None)
    return last_row


def get_last_col(ws: Worksheet, row: int) -> int:
    """Get the index of the last non-empty column in an Excel spreadsheet row."""
    last_col = len(ws[row])
    # make sure we get the first non-empty column from the right
    last_col -= next(i for i, x in enumerate(reversed(ws[row])) if x.value is not None)
    return last_col


def get_as_of_date(date: datetime = datetime.now()) -> str:
    """Get the as-of label for use in column headers,
    defined as the last day of the previous quarter.
    """
    current_month = date.month
    # Jan - Mar
    if current_month <= 3:
        end_date = "12/31"
        year = date.year - 1
    # Apr - Jun
    elif current_month <= 6:
        end_date = "3/31"
        year = date.year
    # Jul - Sep
    elif current_month <= 9:
        end_date = "6/30"
        year = date.year
    # Oct - Dec
    else:
        end_date = "9/30"
        year = date.year
    # return as-of date in MM/DD/YY format
    return f"as of {end_date}/{year%100}"


def df_to_excel(df: pd.DataFrame, ws: Worksheet) -> Worksheet:
    """Puts dataframe into Excel worksheet, with data starting at row 5."""
    # convert to rows for use in spreadsheet
    rows = dataframe_to_rows(df, index=False, header=False)
    # add rows to spreadsheet, starting at row 5
    for r_col_id, row in enumerate(rows, 1):
        for c_col_id, value in enumerate(row, 1):
            ws.cell(row=r_col_id + 4, column=c_col_id, value=value)
    return ws


def create_excel_output(rpt_type: str) -> Workbook:
    """Create Excel output for a report.

    Returns a Workbook, for direct download or archiving as needed.
    """

    # UL and Master reports have extra columns, so use a different template
    if rpt_type in ("master", "ul"):
        template_file = path.join(BASE_DIR, "ge/ge_template_ul.xlsx")
    else:
        template_file = path.join(BASE_DIR, "ge/ge_template.xlsx")
    wb = load_workbook(template_file)

    if rpt_type == "master":
        # only one sheet in master report, so remove the other and rename
        gifts = wb["Gifts"]
        wb.remove_sheet(gifts)
        ws = wb["Endowments"]
        ws.title = "G&E"
        # clear label in template
        ws["A1"] = ""

        # get all data from LibraryData table as a dataframe, excluding SPARE ROWs
        df = pd.DataFrame.from_records(
            LibraryData.objects.filter(~Q(unit="SPARE ROW")).values()
        )
        # get correct cols in correct order
        master_cols = [
            "unit",
            "home_unit_dept",
            "fund_title",
            "fund_type",
            "reg_fdn",
            "fund_manager",
            "ucop_fdn_no",
            "fau_fund_no",
            "fau_account",
            "fau_cost_center",
            "fau_fund",
            "ytd_appropriation",
            "ytd_expenditure",
            "commitments",
            "operating_balance",
            "max_mtf_trf_amt",
            "total_balance",
            "mtf_authority",
            "projected_annual_income",
            "fund_purpose",
            "fund_restriction",
            "notes",
            "lbs_notes",
        ]
        df = df[master_cols]

        ws = df_to_excel(df, ws)

        # add correct cell formatting
        for col in ("L", "M", "N", "O", "P", "Q", "S"):
            for row in range(5, len(ws[col]) + 1):
                # Excel "format code" for Accounting, 2 decimal places, $, comma separator
                ws[
                    f"{col}{row}"
                ].number_format = (
                    """_($* #,##0.00_);_($* (#,##0.00);_($* " - "??_);_(@_)"""
                )
        # add filters on all cols
        filters = ws.auto_filter
        last_col = get_column_letter(get_last_col(ws, 5))
        last_row = get_last_row(ws, "A")
        filters.ref = f"A4:{last_col}{last_row}"

        # add as-of dates to L-O balance cols, max MTF col, and projected annual income
        as_of = get_as_of_date()
        ws["L3"] = as_of
        ws["P3"] = as_of
        ws["S3"] = as_of

    else:
        # map each report type to list of strings needed for query
        rpt_query_dict = {
            "archives": ["Archives"],
            "arts": ["Arts"],
            "biomed": ["Biomed"],
            "digilib": ["DigiLib", "Digital Library"],
            "eal": ["EAL"],
            "ftva": ["FTVA"],
            "hsc": ["History & SC Sciences"],
            "hssd": ["HSSD", "SSHD"],
            "ias": ["Int'l Studies"],
            "lhr": ["LHR"],
            "lsc": ["LSC"],
            "management": ["Management"],
            "music": ["Music"],
            "oh": ["Oral History"],
            "pa": ["Performing Arts"],
            "powell": ["Powell"],
            "preservation": ["Preservation"],
            "sel": ["SEL"],
            "ul": ["UL"],
            "aul_benedetti": ["Benedetti"],
            "aul_consales": ["Consales"],
            "aul_grappone": ["Grappone"],
        }

        # some reports require multiple queries, so start with a blank queryset
        # and OR them together
        endowments_qset = LibraryData.objects.none()
        gifts_qset = LibraryData.objects.none()

        # AUL reports require fuzzy matching on unit and home_unit_dept
        if rpt_type in ["aul_benedetti", "aul_consales", "aul_grappone"]:
            endowments_qset = LibraryData.objects.filter(
                unit__icontains=rpt_query_dict[rpt_type][0]
            ).filter(fund_type="Endowment").order_by(
                "fau_fund_no"
            ) | LibraryData.objects.filter(
                home_unit_dept__icontains=rpt_query_dict[rpt_type][0]
            ).filter(
                fund_type="Endowment"
            ).order_by(
                "fau_fund_no"
            )
            gifts_qset = LibraryData.objects.filter(
                unit__icontains=rpt_query_dict[rpt_type][0]
            ).filter(fund_type="Current Expenditure").order_by(
                "fau_fund_no"
            ) | LibraryData.objects.filter(
                home_unit_dept__icontains=rpt_query_dict[rpt_type][0]
            ).filter(
                fund_type="Current Expenditure"
            ).order_by(
                "fau_fund_no"
            )

        else:
            for query_str in rpt_query_dict[rpt_type]:
                endowments_qset |= (
                    LibraryData.objects.filter(unit=query_str)
                    .filter(fund_type="Endowment")
                    .order_by("fau_fund_no")
                )
                gifts_qset |= (
                    LibraryData.objects.filter(unit=query_str)
                    .filter(fund_type="Current Expenditure")
                    .order_by("fau_fund_no")
                )

        endowments_df = pd.DataFrame.from_records(endowments_qset.values())
        gifts_df = pd.DataFrame.from_records(gifts_qset.values())

        # basic cols for endowments reports
        endowments_cols = [
            "unit",
            "home_unit_dept",
            "fund_title",
            "fund_type",
            "reg_fdn",
            "fund_manager",
            "ucop_fdn_no",
            "fau_fund_no",
            "fau_account",
            "fau_cost_center",
            "fau_fund",
            "ytd_appropriation",
            "ytd_expenditure",
            "commitments",
            "operating_balance",
            "mtf_authority",
            "projected_annual_income",
            "fund_purpose",
            "fund_restriction",
            "notes",
            "lbs_notes",
        ]
        # extra cols for UL report
        if rpt_type == "ul":
            endowments_cols.insert(15, "max_mtf_trf_amt")
            endowments_cols.insert(16, "total_balance")

        if not endowments_df.empty:
            endowments_df = endowments_df[endowments_cols]

            # if there are no fund restrictions, remove that column
            if all(endowments_df["fund_restriction"].isin([""])):
                endowments_df.drop(columns=["fund_restriction"], inplace=True)
                # remove column from Excel template - col U for UL, S for others
                if rpt_type == "ul":
                    wb["Endowments"].delete_cols(21)
                else:
                    wb["Endowments"].delete_cols(19)

        # basic cols for gifts reports
        gifts_cols = [
            "unit",
            "home_unit_dept",
            "fund_title",
            "fund_type",
            "reg_fdn",
            "fund_manager",
            "ucop_fdn_no",
            "fau_fund_no",
            "fau_account",
            "fau_cost_center",
            "fau_fund",
            "ytd_appropriation",
            "ytd_expenditure",
            "commitments",
            "operating_balance",
            "mtf_authority",
            "fund_purpose",
            "fund_restriction",
            "notes",
        ]
        # extra cols for UL report
        if rpt_type == "ul":
            gifts_cols.insert(15, "max_mtf_trf_amt")
            gifts_cols.insert(16, "total_balance")

        if not gifts_df.empty:
            gifts_df = gifts_df[gifts_cols]

            # if there are no fund restrictions, remove that column
            if all(gifts_df["fund_restriction"].isin([""])):
                gifts_df.drop(columns=["fund_restriction"], inplace=True)
                # remove column from Excel template - col T for UL, R for others
                if rpt_type == "ul":
                    wb["Gifts"].delete_cols(20)
                else:
                    wb["Gifts"].delete_cols(18)

        # put data into Excel worksheets
        gifts_ws = wb["Gifts"]
        endowments_ws = wb["Endowments"]
        gifts_ws = df_to_excel(gifts_df, gifts_ws)
        endowments_ws = df_to_excel(endowments_df, endowments_ws)

        # add totals and formatting for money columns
        gifts_money_cols = ["L", "M", "N", "O"]
        endowments_money_cols = ["L", "M", "N", "O", "Q"]
        if rpt_type == "ul":
            gifts_money_cols.extend(["P", "Q"])
            endowments_money_cols.extend(["P", "S"])

        for col in gifts_money_cols:
            sum_col(gifts_ws, col)
            for row in range(5, len(gifts_ws[col]) + 1):
                # Excel "format code" for Accounting, 2 decimal places, $, comma separator
                gifts_ws[
                    f"{col}{row}"
                ].number_format = (
                    """_($* #,##0.00_);_($* (#,##0.00);_($* " - "??_);_(@_)"""
                )

        for col in endowments_money_cols:
            sum_col(endowments_ws, col)
            for row in range(5, len(endowments_ws[col]) + 1):
                # Excel "format code" for Accounting, 2 decimal places, $, comma separator
                endowments_ws[
                    f"{col}{row}"
                ].number_format = (
                    """_($* #,##0.00_);_($* (#,##0.00);_($* " - "??_);_(@_)"""
                )

        # set filters on all columns with data
        endowments_filters = endowments_ws.auto_filter
        last_endowment_col = get_column_letter(get_last_col(endowments_ws, 5))
        last_endowment_row = get_last_row(endowments_ws, "A")
        endowments_filters.ref = f"A4:{last_endowment_col}{last_endowment_row}"

        gifts_filters = gifts_ws.auto_filter
        last_gifts_col = get_column_letter(get_last_col(gifts_ws, 5))
        last_gifts_row = get_last_row(gifts_ws, "A")
        gifts_filters.ref = f"A4:{last_gifts_col}{last_gifts_row}"

        # add as-of dates
        as_of = get_as_of_date()
        # L3 is always the start of the 4 common financial cols
        endowments_ws["L3"] = as_of
        gifts_ws["L3"] = as_of
        if rpt_type == "ul":
            # UL has extra MTF col on both sheets (P), and one other extra col
            # that pushes the Projected Annual Income col to S
            gifts_ws["P3"] = as_of
            endowments_ws["P3"] = as_of
            endowments_ws["S3"] = as_of
        else:
            # Projected Annual Income col is Q on non-UL endowments reports
            endowments_ws["Q3"] = as_of

        # add border formatting to reports
        # row 2 contains top of column headers, and is sometimes merged with 3
        # so we count row 2, but apply the border to row 3
        last_border_col_endowments = get_last_col(endowments_ws, 2)
        # last col (LBS Notes) is excluded from border formatting, and Excel cols are 1-indexed
        # so we use range(1, last_border_col) to get all but the last col
        for header_cell_index in range(1, last_border_col_endowments):
            current_cell = endowments_ws[f"{get_column_letter(header_cell_index)}3"]
            current_cell.border = Border(
                bottom=Side(border_style="medium"),
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
            )
        # Last col before LBS notes needs medium border on right and bottom
        endowments_ws[
            f"{get_column_letter(last_border_col_endowments - 1)}3"
        ].border = Border(
            right=Side(border_style="medium"), bottom=Side(border_style="medium")
        )

        last_border_col_gifts = get_last_col(gifts_ws, 2)
        for header_cell_index in range(1, last_border_col_gifts):
            current_cell = gifts_ws[f"{get_column_letter(header_cell_index)}3"]
            current_cell.border = Border(
                bottom=Side(border_style="medium"),
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
            )
        # Last col before LBS notes needs medium border on right and bottom
        gifts_ws[f"{get_column_letter(last_border_col_gifts - 1)}3"].border = Border(
            right=Side(border_style="medium"), bottom=Side(border_style="medium")
        )

    return wb


def get_bytes_from_workbook(workbook: Workbook) -> bytes:
    """Convert openpyxl workbook into bytes for serving via HTTP."""
    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
    return stream


def download_excel_file(rpt_type: str) -> HttpResponse:
    """Get Excel file via HTTP response."""
    workbook = create_excel_output(rpt_type)

    stream = get_bytes_from_workbook(workbook)

    response = HttpResponse(
        content=stream,
        content_type="application/ms-excel",
    )
    response[
        "Content-Disposition"
    ] = f'attachment; filename={rpt_type}-Report-{datetime.now().strftime("%Y%m%d%H%M")}.xlsx'

    return response


def download_zip_file() -> HttpResponse:
    """Get zip file containing all Excel reports, via HTTP response."""

    # Use the same timestamp for all reports and for zip file.
    timestamp = datetime.now().strftime("%Y%m%d%H%M")
    response = HttpResponse(content_type="application/zip")
    zip_filename = f"ge_reports-{timestamp}.zip"
    zip_file = zipfile.ZipFile(response, "w")

    # Get list of reports from ReportForm.
    report_types = [choice[0] for choice in ReportForm().fields["report_type"].choices]
    # Get Excel workbook for each, convert to bytes, and add to zip file.
    for report_type in report_types:
        workbook = create_excel_output(report_type)
        excel_filename = f"{report_type}-Report-{timestamp}.xlsx"
        stream = get_bytes_from_workbook(workbook)
        zip_file.writestr(excel_filename, stream)

    # Attach it to the response and return it.
    response["Content-Disposition"] = f"attachment; filename={zip_filename}"
    return response
